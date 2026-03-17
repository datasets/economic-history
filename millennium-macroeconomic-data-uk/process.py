"""
Process Bank of England "A Millennium of Macroeconomic Data for the UK" xlsx.

Source: https://www.bankofengland.co.uk/statistics/research-datasets
Version 3.1 (updated to 2016, corrections Dec 2025)

Extracts three headline series:
  - data/annual.csv     — A1. Headline Annual Series (1086-2016)
  - data/quarterly.csv  — Q1. Quarterly Headline Series
  - data/monthly.csv    — M1. Monthly Headline Series
"""
import csv
import re
from collections import defaultdict

import openpyxl

EXCEL_FILE = "millennium-of-macroeconomic-data.xlsx"

# ── Helpers ───────────────────────────────────────────────────────────────────

def safe_float(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        f = float(val)
        import math
        return None if math.isnan(f) or math.isinf(f) else f
    if isinstance(val, str):
        s = val.strip().upper()
        if s in ("#N/A", "N/A", "", "NA", "#VALUE!", "#REF!", "#DIV/0!"):
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def slugify(text):
    """Convert description text to a machine-readable slug."""
    if not text:
        return "unknown"
    s = str(text).strip().lower()
    # Replace common symbols
    s = s.replace("£", "gbp").replace("$", "usd").replace("%", "pct")
    s = s.replace("(", "").replace(")", "").replace("/", "-")
    # Keep only alphanumeric and hyphens
    s = re.sub(r"[^a-z0-9]+", "-", s)
    s = s.strip("-")
    # Truncate at ~60 chars on a word boundary
    if len(s) > 60:
        s = s[:60].rsplit("-", 1)[0]
    return s


def fill_forward(section_row, total_cols):
    """
    Given a row of section header values (sparse), fill-forward sections
    so every column index maps to a section name.
    """
    sections = {}
    current = "General"
    for j in range(total_cols):
        v = section_row[j] if j < len(section_row) else None
        if v and str(v).strip() not in ("", "Section", "Back to front page"):
            current = str(v).strip()
        sections[j] = current
    return sections


def build_var_meta_annual(rows):
    """
    Build variable metadata from A1. Headline series header rows.

    Header layout (0-indexed rows):
      0: title row
      1: "Back to front page" row
      2: Section row
      3: Description row (row 4 in 1-indexed)
      4: Worksheet row
      5: Units row
      6: Documentation row
      7+: data rows starting from 1086

    Data columns start at col index 1 (col 2 in 1-indexed, since col 1 = year).
    """
    # rows 0-6 are header rows in 0-indexed
    section_row_vals = [cell.value for cell in rows[2]]
    desc_row_vals = [cell.value for cell in rows[3]]
    unit_row_vals = [cell.value for cell in rows[5]]

    total_cols = len(section_row_vals)
    sections = fill_forward(section_row_vals, total_cols)

    metas = []  # list of (col_idx, variable_id, description, section, unit)
    seen_slugs = defaultdict(int)

    for j in range(1, total_cols):  # skip col 0 = year
        desc = desc_row_vals[j] if j < len(desc_row_vals) else None
        unit = unit_row_vals[j] if j < len(unit_row_vals) else None
        if not desc:
            continue

        desc = str(desc).strip()
        unit = str(unit).strip() if unit else ""
        section = sections[j]

        slug = slugify(desc)
        seen_slugs[slug] += 1
        if seen_slugs[slug] > 1:
            slug = f"{slug}-{seen_slugs[slug]}"

        metas.append((j, slug, desc, section, unit))

    return metas


def build_var_meta_quarterly(rows):
    """
    Q1. Qrtly headline series header layout (0-indexed):
      0: back + title
      1: empty
      2: Section row
      3: Description row
      4: Worksheet
      5: Units
      6: Documentation
      7+: data (col 0=year, col 1=quarter, data from col 2)
    """
    section_row_vals = [cell.value for cell in rows[2]]
    desc_row_vals = [cell.value for cell in rows[3]]
    unit_row_vals = [cell.value for cell in rows[5]]

    total_cols = len(section_row_vals)
    sections = fill_forward(section_row_vals, total_cols)

    metas = []
    seen_slugs = defaultdict(int)

    for j in range(2, total_cols):  # skip col 0=year, col 1=quarter
        desc = desc_row_vals[j] if j < len(desc_row_vals) else None
        unit = unit_row_vals[j] if j < len(unit_row_vals) else None
        if not desc:
            continue

        desc = str(desc).strip()
        unit = str(unit).strip() if unit else ""
        section = sections[j]

        slug = slugify(desc)
        seen_slugs[slug] += 1
        if seen_slugs[slug] > 1:
            slug = f"{slug}-{seen_slugs[slug]}"

        metas.append((j, slug, desc, section, unit))

    return metas


def build_var_meta_monthly(rows):
    """
    M1. Mthly headline series header layout (0-indexed):
      0: back + title
      1: empty
      2: empty
      3: Section row (row 4)
      4: empty
      5: Description row (row 6)
      6: Worksheet
      7: Units
      8: Seasonality
      9: Notes
      10+: data (col 0=year, col 1=month, data from col 2)
    """
    section_row_vals = [cell.value for cell in rows[3]]
    desc_row_vals = [cell.value for cell in rows[5]]
    unit_row_vals = [cell.value for cell in rows[7]]

    total_cols = max(len(section_row_vals), len(desc_row_vals), len(unit_row_vals))
    sections = fill_forward(section_row_vals, total_cols)

    metas = []
    seen_slugs = defaultdict(int)

    for j in range(2, total_cols):  # skip col 0=year, col 1=month
        desc = desc_row_vals[j] if j < len(desc_row_vals) else None
        unit = unit_row_vals[j] if j < len(unit_row_vals) else None
        if not desc:
            continue

        desc = str(desc).strip()
        unit = str(unit).strip() if unit else ""
        section = sections[j]

        slug = slugify(desc)
        seen_slugs[slug] += 1
        if seen_slugs[slug] > 1:
            slug = f"{slug}-{seen_slugs[slug]}"

        metas.append((j, slug, desc, section, unit))

    return metas


# ── Extract annual data ───────────────────────────────────────────────────────

def extract_annual(wb):
    ws = wb["A1. Headline series"]
    rows = list(ws.iter_rows(values_only=False))

    metas = build_var_meta_annual(rows)
    print(f"  Annual: {len(metas)} variables defined")

    out = []
    current_year = None

    for row in rows[7:]:  # data starts at row index 7 (row 8, 1-indexed)
        vals = [cell.value for cell in row]
        year_val = vals[0]

        if isinstance(year_val, (int, float)):
            current_year = int(year_val)
        elif year_val is not None:
            continue  # skip non-numeric year values

        if current_year is None:
            continue

        for (j, slug, desc, section, unit) in metas:
            v = safe_float(vals[j]) if j < len(vals) else None
            if v is None:
                continue
            out.append({
                "year": current_year,
                "variable_id": slug,
                "variable": desc,
                "section": section,
                "unit": unit,
                "value": v,
            })

    return out


# ── Extract quarterly data ────────────────────────────────────────────────────

def extract_quarterly(wb):
    ws = wb["Q1. Qrtly headline series"]
    rows = list(ws.iter_rows(values_only=False))

    metas = build_var_meta_quarterly(rows)
    print(f"  Quarterly: {len(metas)} variables defined")

    out = []
    current_year = None

    for row in rows[7:]:
        vals = [cell.value for cell in row]
        year_val = vals[0]
        quarter_val = vals[1] if len(vals) > 1 else None

        if isinstance(year_val, (int, float)):
            current_year = int(year_val)

        if current_year is None:
            continue
        if not quarter_val or not str(quarter_val).strip().startswith("Q"):
            continue

        quarter = str(quarter_val).strip()
        qnum = int(quarter[1])
        period = f"{current_year}-{quarter}"

        for (j, slug, desc, section, unit) in metas:
            v = safe_float(vals[j]) if j < len(vals) else None
            if v is None:
                continue
            out.append({
                "period": period,
                "year": current_year,
                "quarter": qnum,
                "variable_id": slug,
                "variable": desc,
                "section": section,
                "unit": unit,
                "value": v,
            })

    return out


# ── Extract monthly data ──────────────────────────────────────────────────────

MONTH_MAP = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}


def extract_monthly(wb):
    ws = wb["M1. Mthly headline series"]
    rows = list(ws.iter_rows(values_only=False))

    metas = build_var_meta_monthly(rows)
    print(f"  Monthly: {len(metas)} variables defined")

    out = []
    current_year = None

    for row in rows[10:]:
        vals = [cell.value for cell in row]
        year_val = vals[0]
        month_val = vals[1] if len(vals) > 1 else None

        if isinstance(year_val, (int, float)):
            current_year = int(year_val)

        if current_year is None:
            continue
        if not month_val:
            continue

        month_str = str(month_val).strip().lower()
        month_num = MONTH_MAP.get(month_str[:3])
        if month_num is None:
            continue

        period = f"{current_year}-{month_num:02d}"

        for (j, slug, desc, section, unit) in metas:
            v = safe_float(vals[j]) if j < len(vals) else None
            if v is None:
                continue
            out.append({
                "period": period,
                "year": current_year,
                "month": month_num,
                "variable_id": slug,
                "variable": desc,
                "section": section,
                "unit": unit,
                "value": v,
            })

    return out


# ── Write CSVs ────────────────────────────────────────────────────────────────

def write_csv(rows, path, fieldnames):
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k, "") for k in fieldnames})
    print(f"  Wrote {len(rows):,} rows → {path}")


# ── Main ──────────────────────────────────────────────────────────────────────

import warnings
warnings.filterwarnings("ignore")

print("Loading workbook (this may take a moment for 26MB xlsx)...")
wb = openpyxl.load_workbook(EXCEL_FILE, read_only=False, data_only=True)

print("\nExtracting annual headline series (A1)...")
annual_rows = extract_annual(wb)
annual_rows.sort(key=lambda r: (r["variable_id"], r["year"]))
write_csv(
    annual_rows,
    "data/annual.csv",
    ["year", "variable_id", "variable", "section", "unit", "value"],
)

print("\nExtracting quarterly headline series (Q1)...")
quarterly_rows = extract_quarterly(wb)
quarterly_rows.sort(key=lambda r: (r["variable_id"], r["year"], r["quarter"]))
write_csv(
    quarterly_rows,
    "data/quarterly.csv",
    ["period", "year", "quarter", "variable_id", "variable", "section", "unit", "value"],
)

print("\nExtracting monthly headline series (M1)...")
monthly_rows = extract_monthly(wb)
monthly_rows.sort(key=lambda r: (r["variable_id"], r["period"]))
write_csv(
    monthly_rows,
    "data/monthly.csv",
    ["period", "year", "month", "variable_id", "variable", "section", "unit", "value"],
)

wb.close()

# Summary
print("\n=== Summary ===")
from collections import Counter
by_var = Counter(r["variable_id"] for r in annual_rows)
annual_vars = len(by_var)
if annual_rows:
    years = [r["year"] for r in annual_rows]
    print(f"  Annual:    {len(annual_rows):,} rows, {annual_vars} variables, years {min(years)}-{max(years)}")

if quarterly_rows:
    yrs = [r["year"] for r in quarterly_rows]
    qvars = len(set(r["variable_id"] for r in quarterly_rows))
    print(f"  Quarterly: {len(quarterly_rows):,} rows, {qvars} variables, years {min(yrs)}-{max(yrs)}")

if monthly_rows:
    yrs = [r["year"] for r in monthly_rows]
    mvars = len(set(r["variable_id"] for r in monthly_rows))
    print(f"  Monthly:   {len(monthly_rows):,} rows, {mvars} variables, years {min(yrs)}-{max(yrs)}")
