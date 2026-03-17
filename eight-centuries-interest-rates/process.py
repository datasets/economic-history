"""
process.py — Eight Centuries of Global Real Interest Rates

Source:
  Paul Schmelzing (2020), "Eight Centuries of Global Real Interest Rates, R-G,
  and the 'Suprasecular' Decline, 1311–2018", Bank of England Staff Working
  Paper No. 845.

  Data xlsx from GitHub issue:
  https://github.com/datasets/awesome-data/files/6604635/
    eight-centuries-of-global-real-interest-rates-r-g-and-the-suprasecular-decline-1311-2018-data.xlsx

Outputs:
  data/headline-rates.csv   — Annual global real rates (7yr-avg and not-avg), 1311–2018
  data/country-rates.csv    — Country-level real rates (not averaged), 1310–2018
"""

import csv
import io
import urllib.request

import openpyxl

XLSX_URL = (
    "https://github.com/datasets/awesome-data/files/6604635/"
    "eight-centuries-of-global-real-interest-rates-r-g-and-the-suprasecular"
    "-decline-1311-2018-data.xlsx"
)


def download_bytes(url):
    with urllib.request.urlopen(url) as r:
        return r.read()


def parse_headline(wb):
    """Sheet II: Headline series — global real rates, 7yr-avg and not-avg."""
    ws = wb["II. Headline series"]
    rows = list(ws.iter_rows(values_only=True))

    # Column mapping (0-indexed):
    #   7-year averaged block (cols 2-7):
    #     2: global_nominal_7y, 3: global_inflation_7y, 4: global_real_7y
    #     5: safe_nominal_7y,   6: safe_inflation_7y,   7: safe_real_7y
    #   Not averaged block (cols 10-15):
    #     10: global_nominal, 11: global_inflation, 12: global_real
    #     13: safe_nominal,   14: safe_inflation,   15: safe_real

    records = []
    for row in rows:
        year = row[0]
        if not isinstance(year, (int, float)):
            continue
        year = int(year)

        def val(col):
            v = row[col] if col < len(row) else None
            if v is None or not isinstance(v, (int, float)):
                return ""
            return round(v, 6)

        record = {
            "year": year,
            # 7-year centred average
            "global_nominal_7y": val(2),
            "global_inflation_7y": val(3),
            "global_real_7y": val(4),
            "safe_nominal_7y": val(5),
            "safe_inflation_7y": val(6),
            "safe_real_7y": val(7),
            # Annual (not averaged)
            "global_nominal": val(10),
            "global_inflation": val(11),
            "global_real": val(12),
            "safe_nominal": val(13),
            "safe_inflation": val(14),
            "safe_real": val(15),
        }

        # Skip rows where all rate fields are empty
        if all(record[k] == "" for k in record if k != "year"):
            continue

        records.append(record)

    return records


def parse_country(wb):
    """Sheet IV: Country-level real rates (not averaged), long format.

    Note: The year column (col 0) contains formulas referencing an external
    workbook. The cached values are correct for most rows but the last 6 rows
    (2013-2018) all show 2018 due to stale formula cache. We derive the year
    from the row index instead, which is reliable.
    """
    ws = wb["IV. Country level, 1310-2018"]
    rows = list(ws.iter_rows(values_only=True))

    # Data starts at row index 7 with year 1314.
    DATA_START_ROW = 7
    DATA_START_YEAR = 1314

    # Country columns (0-indexed) in the "Real rates, not averaged" block:
    # col 2: Italy, 3: UK, 4: Holland/Netherlands, 5: Germany, 6: France,
    # 7: United States, 8: Spain, 9: Japan
    countries = {
        2: "Italy",
        3: "UK",
        4: "Netherlands",
        5: "Germany",
        6: "France",
        7: "United States",
        8: "Spain",
        9: "Japan",
    }

    records = []
    for i, row in enumerate(rows):
        if i < DATA_START_ROW:
            continue
        # Derive year from row position (handles stale formula cache at end)
        year = DATA_START_YEAR + (i - DATA_START_ROW)
        if year > 2018:
            break

        for col, country in countries.items():
            v = row[col] if col < len(row) else None
            if v is None or not isinstance(v, (int, float)):
                continue
            records.append({
                "year": year,
                "country": country,
                "real_rate": round(v, 6),
            })

    return records


def write_csv(path, fieldnames, rows):
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    print(f"  Wrote {len(rows)} rows → {path}")


def main():
    print("Downloading Schmelzing (2020) xlsx data...")
    data = download_bytes(XLSX_URL)

    print("Parsing Excel (data_only mode)...")
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)

    print("Extracting headline rates...")
    headline = parse_headline(wb)

    print("Extracting country-level rates...")
    country = parse_country(wb)

    print("Writing CSVs...")
    headline_fields = [
        "year",
        "global_nominal_7y", "global_inflation_7y", "global_real_7y",
        "safe_nominal_7y", "safe_inflation_7y", "safe_real_7y",
        "global_nominal", "global_inflation", "global_real",
        "safe_nominal", "safe_inflation", "safe_real",
    ]
    write_csv("data/headline-rates.csv", headline_fields, headline)

    country_fields = ["year", "country", "real_rate"]
    write_csv("data/country-rates.csv", country_fields, country)

    print("Done.")


if __name__ == "__main__":
    main()
